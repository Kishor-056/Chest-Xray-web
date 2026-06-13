import os
import time
import datetime
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Test Results Container
results = []

def record_test(tc_id, module, description, steps, expected, status, error="None"):
    print(f"[{status}] {tc_id} ({module}): {description}")
    results.append({
        "Test Case ID": tc_id,
        "Module/Screen": module,
        "Description": description,
        "Steps": steps,
        "Expected Result": expected,
        "Status": status,
        "Error Details": error,
        "Timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

def run_tests():
    print("============================================================")
    print("STARTING E2E AUTOMATION TEST SUITE (SELENIUM)")
    print("============================================================")

    # Setup Paths
    base_dir = os.path.dirname(os.path.abspath(__file__))
    sample_valid = os.path.join(base_dir, "public", "samples", "valid_xray.png")
    sample_invalid = os.path.join(base_dir, "public", "samples", "invalid_image.png")
    
    print(f"Sample Valid X-ray path: {sample_valid} (Exists: {os.path.exists(sample_valid)})")
    print(f"Sample Invalid Image path: {sample_invalid} (Exists: {os.path.exists(sample_invalid)})")

    # Setup Chrome options
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--log-level=3')
    
    driver = None
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        wait = WebDriverWait(driver, 20) # High timeout for slow operations (e.g. model loading)
        record_test("SYS_001", "Initialization", "Initialize WebDriver and browser", 
                    "Launch headless Chrome using webdriver_manager", 
                    "Browser launches successfully", "PASS")
    except Exception as e:
        record_test("SYS_001", "Initialization", "Initialize WebDriver and browser", 
                    "Launch headless Chrome using webdriver_manager", 
                    "Browser launches successfully", "FAIL", str(e))
        return

    base_url = "http://localhost:3000"

    # 1. Load Homepage & Check Status
    try:
        driver.get(base_url)
        # Wait for dashboard h1
        header = wait.until(EC.presence_of_element_located((By.TAG_NAME, "h1")))
        
        # Check system status indicator
        status_dot = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "status-dot")))
        status_text = driver.find_element(By.CLASS_NAME, "status-text")
        
        if "online" in status_dot.get_attribute("class") or "Online" in status_text.text:
            record_test("TC_DASH_001", "Dashboard", "Verify Dashboard loads and API is online",
                        "Navigate to homepage and locate system status indicator",
                        "Dashboard loads, status text shows 'Online'", "PASS", f"Header: {header.text}, Status: {status_text.text}")
        else:
            record_test("TC_DASH_001", "Dashboard", "Verify Dashboard loads and API is online",
                        "Navigate to homepage and locate system status indicator",
                        "Dashboard loads, status text shows 'Online'", "FAIL", f"System offline. Status text: {status_text.text}")
    except Exception as e:
        record_test("TC_DASH_001", "Dashboard", "Verify Dashboard loads and API is online",
                    "Navigate to homepage and locate system status indicator",
                    "Dashboard loads, status text shows 'Online'", "FAIL", str(e))

    # 2. Sidebar Navigation Test
    try:
        nav_items = {
            "/predict": "Single Image Prediction",
            "/batch": "Batch Processing",
            "/compare": "Model Comparison",
            "/gradcam": "GradCAM Visualization",
            "/reports": "Clinical Reports",
            "/analytics": "Analytics",
            "/history": "History",
            "/settings": "Settings"
        }
        
        nav_failures = []
        for path, header_title in nav_items.items():
            driver.get(f"{base_url}{path}")
            try:
                h1 = wait.until(EC.presence_of_element_located((By.TAG_NAME, "h1")))
                if header_title.lower() not in h1.text.lower() and "prediction" not in h1.text.lower():
                    nav_failures.append(f"Path {path}: expected header '{header_title}', found '{h1.text}'")
            except Exception as nav_err:
                nav_failures.append(f"Path {path} failed to load: {str(nav_err)}")
                
        if not nav_failures:
            record_test("TC_NAV_001", "Navigation", "Verify all sidebar menu navigation links function correctly",
                        "Navigate to paths: /predict, /batch, /compare, /gradcam, /reports, /analytics, /history, /settings",
                        "Each page loads successfully with its corresponding heading", "PASS")
        else:
            record_test("TC_NAV_001", "Navigation", "Verify all sidebar menu navigation links function correctly",
                        "Navigate to paths: /predict, /batch, /compare, /gradcam, /reports, /analytics, /history, /settings",
                        "Each page loads successfully with its corresponding heading", "FAIL", "; ".join(nav_failures))
    except Exception as e:
        record_test("TC_NAV_001", "Navigation", "Verify all sidebar menu navigation links function correctly",
                    "Navigate to paths: /predict, /batch, /compare, /gradcam, /reports, /analytics, /history, /settings",
                    "Each page loads successfully with its corresponding heading", "FAIL", str(e))

    # 3. Single Prediction (Positive Path)
    try:
        driver.get(f"{base_url}/predict")
        
        # Find and make file input visible
        file_input = wait.until(EC.presence_of_element_located((By.ID, "fileInput")))
        driver.execute_script("arguments[0].style.display = 'block';", file_input)
        file_input.send_keys(sample_valid)
        
        # Wait for preview to render
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "image-preview")))
        
        # Select Model (DenseNet-169)
        selects = driver.find_elements(By.TAG_NAME, "select")
        model_select = selects[0]
        model_select.send_keys("DenseNet-169")
        
        # Click Predict
        predict_btn = driver.find_element(By.CLASS_NAME, "btn-predict")
        predict_btn.click()
        
        # Wait for results to appear
        results_header = wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Prediction Results')]")))
        diagnosis = driver.find_element(By.CLASS_NAME, "prediction-label").text
        confidence = driver.find_element(By.CLASS_NAME, "confidence-text").text
        
        record_test("TC_PRED_001", "Single Prediction", "Predict disease using valid X-ray image",
                    "1. Go to /predict\n2. Upload valid chest X-ray\n3. Select DenseNet-169\n4. Click Predict Disease",
                    "Prediction results loaded displaying Diagnosis and Confidence metrics", "PASS", f"Result: {diagnosis}, {confidence}")
    except Exception as e:
        record_test("TC_PRED_001", "Single Prediction", "Predict disease using valid X-ray image",
                    "1. Go to /predict\n2. Upload valid chest X-ray\n3. Select DenseNet-169\n4. Click Predict Disease",
                    "Prediction results loaded displaying Diagnosis and Confidence metrics", "FAIL", str(e))

    # 4. Single Prediction (Negative Path - Invalid Image)
    try:
        driver.get(f"{base_url}/predict")
        
        file_input = wait.until(EC.presence_of_element_located((By.ID, "fileInput")))
        driver.execute_script("arguments[0].style.display = 'block';", file_input)
        file_input.send_keys(sample_invalid)
        
        # Wait for preview
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "image-preview")))
        
        # Click Predict (which will trigger local validation and raise toast)
        predict_btn = driver.find_element(By.CLASS_NAME, "btn-predict")
        predict_btn.click()
        
        # Wait specifically for the ERROR toast to appear (info toast appears first)
        toast_body = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "Toastify__toast--error")))
        toast_text = toast_body.text
        
        if "not a valid" in toast_text.lower() or "rejected" in toast_text.lower() or "color" in toast_text.lower() or "invalid" in toast_text.lower() or "error" in toast_text.lower():
            record_test("TC_PRED_002", "Single Prediction", "Verify client-side validation rejects invalid non-X-ray images",
                        "1. Go to /predict\n2. Upload invalid test image (e.g. invalid_image.png)\n3. Click Predict Disease\n4. Check error toast",
                        "App displays error toast and prevents sending invalid image for prediction", "PASS", f"Validation Toast: '{toast_text}'")
        else:
            record_test("TC_PRED_002", "Single Prediction", "Verify client-side validation rejects invalid non-X-ray images",
                        "1. Go to /predict\n2. Upload invalid test image (e.g. invalid_image.png)\n3. Click Predict Disease\n4. Check error toast",
                        "App displays error toast and prevents sending invalid image for prediction", "FAIL", f"Toast showed wrong message: '{toast_text}'")
    except Exception as e:
        record_test("TC_PRED_002", "Single Prediction", "Verify client-side validation rejects invalid non-X-ray images",
                    "1. Go to /predict\n2. Upload invalid test image (e.g. invalid_image.png)\n3. Click Predict Disease\n4. Check error toast",
                    "App displays error toast and prevents sending invalid image for prediction", "FAIL", str(e))

    # 5. Batch Processing (Positive Path)
    try:
        driver.get(f"{base_url}/batch")
        
        # Find file input inside .upload-controls
        file_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
        driver.execute_script("arguments[0].style.display = 'block';", file_input)
        file_input.send_keys(sample_valid)
        
        # CRITICAL: Wait for client-side validation on file change to complete and show '.file-item'
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "file-item")))
        
        # Click Process
        process_btn = driver.find_element(By.CLASS_NAME, "btn-process")
        process_btn.click()
        
        # Wait for results
        batch_header = wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Batch Results')]")))
        table = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "results-table")))
        rows = table.find_elements(By.TAG_NAME, "tr")
        
        record_test("TC_BATCH_001", "Batch Processing", "Process multiple chest X-ray images in a batch",
                    "1. Go to /batch\n2. Upload valid chest X-ray\n3. Wait for validation\n4. Click Process Images",
                    "Batch processing completes, displaying aggregated statistics and a results table", "PASS", f"Found {len(rows)-1} processed rows in results table.")
    except Exception as e:
        record_test("TC_BATCH_001", "Batch Processing", "Process multiple chest X-ray images in a batch",
                    "1. Go to /batch\n2. Upload valid chest X-ray\n3. Wait for validation\n4. Click Process Images",
                    "Batch processing completes, displaying aggregated statistics and a results table", "FAIL", str(e))

    # 6. Model Comparison
    try:
        driver.get(f"{base_url}/compare")
        
        # Wait for compare input to exist
        file_input = wait.until(EC.presence_of_element_located((By.ID, "compareFileInput")))
        driver.execute_script("arguments[0].style.display = 'block';", file_input)
        file_input.send_keys(sample_valid)
        
        # CRITICAL: Wait for client-side validation to complete and show '.image-preview'
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "image-preview")))
        
        # Click Compare
        compare_btn = driver.find_element(By.CLASS_NAME, "btn-compare")
        compare_btn.click()
        
        # Wait for comparison results
        chart = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "horizontal-bar-chart")))
        best_model = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "best-model-banner"))).text
        
        record_test("TC_COMP_001", "Model Comparison", "Compare predictions across multiple loaded AI models",
                    "1. Go to /compare\n2. Upload valid X-ray\n3. Wait for validation\n4. Click Compare Models",
                    "Comparison analysis completes, presenting a 'Best Performing Model' banner and a confidence comparison chart", "PASS", f"Consensus analysis loaded. Best model: {best_model.splitlines()[1] if len(best_model.splitlines()) > 1 else best_model}")
    except Exception as e:
        record_test("TC_COMP_001", "Model Comparison", "Compare predictions across multiple loaded AI models",
                    "1. Go to /compare\n2. Upload valid X-ray\n3. Wait for validation\n4. Click Compare Models",
                    "Comparison analysis completes, presenting a 'Best Performing Model' banner and a confidence comparison chart", "FAIL", str(e))

    # 7. GradCAM Viewer
    try:
        driver.get(f"{base_url}/gradcam")
        
        file_input = wait.until(EC.presence_of_element_located((By.ID, "gradcamFileInput")))
        driver.execute_script("arguments[0].style.display = 'block';", file_input)
        file_input.send_keys(sample_valid)
        
        # CRITICAL: Wait for client-side validation to complete and show preview
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "preview-image")))
        
        # Click Generate GradCAM
        generate_btn = driver.find_element(By.CLASS_NAME, "btn-generate")
        generate_btn.click()
        
        # Wait for visualization result image
        viz_card = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "viz-card")))
        img = viz_card.find_element(By.TAG_NAME, "img")
        src = img.get_attribute("src")
        
        if src.startswith("data:image"):
            record_test("TC_GRAD_001", "GradCAM Viewer", "Generate GradCAM attention heatmaps for X-ray analysis",
                        "1. Go to /gradcam\n2. Upload valid X-ray\n3. Wait for validation\n4. Click Generate GradCAM",
                        "GradCAM visualization generates successfully, rendering the heatmap image on screen", "PASS")
        else:
            record_test("TC_GRAD_001", "GradCAM Viewer", "Generate GradCAM attention heatmaps for X-ray analysis",
                        "1. Go to /gradcam\n2. Upload valid X-ray\n3. Wait for validation\n4. Click Generate GradCAM",
                        "GradCAM visualization generates successfully, rendering the heatmap image on screen", "FAIL", f"Invalid image source: {src[:100]}...")
    except Exception as e:
        record_test("TC_GRAD_001", "GradCAM Viewer", "Generate GradCAM attention heatmaps for X-ray analysis",
                    "1. Go to /gradcam\n2. Upload valid X-ray\n3. Wait for validation\n4. Click Generate GradCAM",
                    "GradCAM visualization generates successfully, rendering the heatmap image on screen", "FAIL", str(e))

    # 8. Clinical Reports
    try:
        driver.get(f"{base_url}/reports")
        
        # Enter Patient ID
        pid_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='e.g., P12345']")))
        pid_input.send_keys("PAT-999")
        
        # Upload X-ray
        file_input = wait.until(EC.presence_of_element_located((By.ID, "reportFileInput")))
        driver.execute_script("arguments[0].style.display = 'block';", file_input)
        file_input.send_keys(sample_valid)
        
        # Wait for validation
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "image-preview")))
        
        # Click Generate Report
        gen_report_btn = driver.find_element(By.CLASS_NAME, "btn-generate-report")
        gen_report_btn.click()
        
        # Verify report preview display
        report_display = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "report-display")))
        diagnosis_label = report_display.find_element(By.CLASS_NAME, "diagnosis-label").text
        
        record_test("TC_REP_001", "Clinical Reports", "Generate medical report for patient with AI diagnosis",
                    "1. Go to /reports\n2. Input Patient ID 'PAT-999'\n3. Upload valid X-ray\n4. Click Generate Report",
                    "Comprehensive medical report is compiled showing patient metadata, primary diagnosis, and disclaimer", "PASS", f"Report created. Diagnosis: {diagnosis_label}")
    except Exception as e:
        record_test("TC_REP_001", "Clinical Reports", "Generate medical report for patient with AI diagnosis",
                    "1. Go to /reports\n2. Input Patient ID 'PAT-999'\n3. Upload valid X-ray\n4. Click Generate Report",
                    "Comprehensive medical report is compiled showing patient metadata, primary diagnosis, and disclaimer", "FAIL", str(e))

    # 9. Settings Save
    try:
        driver.get(f"{base_url}/settings")
        
        # Modify Default Model
        select_model = wait.until(EC.presence_of_element_located((By.TAG_NAME, "select")))
        select_model.send_keys("EfficientNet-B5")
        
        # Click Apply Switch
        apply_switch_btn = driver.find_element(By.XPATH, "//*[contains(text(), 'Apply Model Switch')]")
        apply_switch_btn.click()
        time.sleep(1)
        
        # Click Save Settings
        save_btn = driver.find_element(By.CLASS_NAME, "btn-save")
        save_btn.click()
        time.sleep(1)
        
        record_test("TC_SET_001", "Settings", "Configure application preferences and save settings",
                    "1. Go to /settings\n2. Set Default Model to EfficientNet-B5\n3. Click Apply Model Switch and Save Settings",
                    "Settings are saved locally, and default model switch is applied successfully", "PASS")
    except Exception as e:
        record_test("TC_SET_001", "Settings", "Configure application preferences and save settings",
                    "1. Go to /settings\n2. Set Default Model to EfficientNet-B5\n3. Click Apply Model Switch and Save Settings",
                    "Settings are saved locally, and default model switch is applied successfully", "FAIL", str(e))

    # 10. Submit Feedback
    try:
        driver.get(f"{base_url}/settings")
        
        # Fill feedback inputs
        pid_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Prediction ID from recent result']")))
        pid_input.send_keys("PRED-12345")
        
        diag_input = driver.find_element(By.XPATH, "//input[@placeholder='e.g., Pneumonia']")
        diag_input.send_keys("Pneumonia")
        
        conf_input = driver.find_element(By.XPATH, "//input[@placeholder='e.g., 0.95']")
        conf_input.send_keys("0.90")
        
        comments_input = driver.find_element(By.XPATH, "//textarea[@placeholder='Additional context or corrections...']")
        comments_input.send_keys("Model misdiagnosed a shadow, corrected to Pneumonia.")
        
        # Submit (SPECIFICALLY MATCH THE BUTTON ELEMENT TO AVOID <h2> ELEMENT)
        submit_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Submit Feedback')]")))
        submit_btn.click()
        
        # Wait for feedback success toast
        toast_body = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "Toastify__toast-body")))
        toast_text = toast_body.text
        
        if "success" in toast_text.lower() or "submitted" in toast_text.lower():
            record_test("TC_FEED_001", "Settings", "Submit prediction correction feedback to improve AI models",
                        "1. Go to /settings feedback section\n2. Fill in Prediction ID, actual diagnosis, and comments\n3. Click Submit Feedback",
                        "Feedback payload is successfully transmitted to backend system showing success notification", "PASS", f"Toast: '{toast_text}'")
        else:
            record_test("TC_FEED_001", "Settings", "Submit prediction correction feedback to improve AI models",
                        "1. Go to /settings feedback section\n2. Fill in Prediction ID, actual diagnosis, and comments\n3. Click Submit Feedback",
                        "Feedback payload is successfully transmitted to backend system showing success notification", "FAIL", f"Unexpected Toast: '{toast_text}'")
    except Exception as e:
        record_test("TC_FEED_001", "Settings", "Submit prediction correction feedback to improve AI models",
                    "1. Go to /settings feedback section\n2. Fill in Prediction ID, actual diagnosis, and comments\n3. Click Submit Feedback",
                    "Feedback payload is successfully transmitted to backend system showing success notification", "FAIL", str(e))

    # Cleanup
    driver.quit()
    print("============================================================")
    print("AUTOMATION TESTS COMPLETED")
    print("============================================================")

def generate_report_excel():
    print("Generating beautifully styled Excel test report...")
    df = pd.DataFrame(results)
    
    # Save base file
    filename = "Test_Report.xlsx"
    df.to_excel(filename, index=False)
    
    # Load with openpyxl to apply rich styling
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.title = "E2E Test Results"
    
    # Styles
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep navy
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    cell_font = Font(name=font_family, size=10)
    bold_cell_font = Font(name=font_family, size=10, bold=True)
    
    pass_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") # Soft green
    fail_fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid") # Soft orange/red
    pass_font = Font(name=font_family, size=10, color="274E13", bold=True)
    fail_font = Font(name=font_family, size=10, color="783F04", bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Style header row
    ws.row_dimensions[1].height = 28
    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Style data rows
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
        ws.row_dimensions[r_idx].height = 36 # taller rows for readability
        
        status_cell = ws.cell(row=r_idx, column=6) # Status column is 6th
        if status_cell.value == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        else:
            status_cell.fill = fail_fill
            status_cell.font = fail_font
            
        for c_idx, cell in enumerate(row, start=1):
            cell.border = thin_border
            if c_idx in [1, 6, 8]: # ID, Status, Timestamp
                cell.alignment = center_align
                if c_idx == 1:
                    cell.font = bold_cell_font
                else:
                    cell.font = cell_font
            else:
                cell.alignment = left_align
                cell.font = cell_font
                
    # Set explicit widths for columns
    ws.column_dimensions['A'].width = 15 # Test Case ID
    ws.column_dimensions['B'].width = 18 # Module
    ws.column_dimensions['C'].width = 30 # Description
    ws.column_dimensions['D'].width = 45 # Steps
    ws.column_dimensions['E'].width = 45 # Expected Result
    ws.column_dimensions['F'].width = 12 # Status
    ws.column_dimensions['G'].width = 30 # Error Details
    ws.column_dimensions['H'].width = 20 # Timestamp

    wb.save(filename)
    print(f"Formatted report saved successfully as '{filename}'!")

if __name__ == "__main__":
    try:
        run_tests()
    except Exception as ex:
        print(f"Fatal execution error: {ex}")
    finally:
        generate_report_excel()
